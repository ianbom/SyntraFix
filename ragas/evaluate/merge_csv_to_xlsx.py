import csv
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


BASE_DIR = Path(__file__).resolve().parent
OUTPUT = BASE_DIR / "final-evaluate-updated.xlsx"
FILES = [
    "sample-bgem3-samples-001-005.csv",
    "sample-bgem3-samples-006-010.csv",
    "sample-bgem3-samples-011-015.csv",
    "sample-bgem3-samples-016-020.csv",
    "sample-bgem3-samples-021-025.csv",
    "sample-bgem3-samples-026-030.csv",
    "sample-bgem3-samples-031-035.csv",
    "sample-bgem3-samples-036-040.csv",
    "sample-bgem3-samples-041-045.csv",
    "sample-bgem3-samples-046-050.csv",
    "sample-bgem3-samples-051-055.csv",
    "sample-bgem3-samples-056-059.csv",
]


def read_csv(path: Path):
    with path.open("r", newline="", encoding="utf-8-sig") as handle:
        reader = csv.reader(handle)
        header = next(reader)
        rows = list(reader)
    return header, rows


def build_workbook():
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Evaluation"

    expected_header = None
    output_rows = []
    source_counts = []

    for name in FILES:
        header, rows = read_csv(BASE_DIR / name)
        if expected_header is None:
            expected_header = header
            sheet.append(expected_header)
        elif header != expected_header:
            raise ValueError(f"Header mismatch in {name}")

        output_rows.extend(rows)
        source_counts.append((name, len(rows)))

    for row in output_rows:
        sheet.append(row)

    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions

    widths = {
        "A": 28,
        "B": 34,
        "C": 70,
        "D": 55,
        "E": 55,
        "F": 14,
        "G": 18,
        "H": 18,
        "I": 16,
    }
    for column, width in widths.items():
        sheet.column_dimensions[column].width = width

    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    sheet.row_dimensions[1].height = 28

    workbook.save(OUTPUT)
    return len(output_rows), source_counts


def verify_workbook(expected_rows: int):
    workbook = load_workbook(OUTPUT, read_only=True, data_only=False)
    sheet = workbook["Evaluation"]
    data_rows = sheet.max_row - 1
    if data_rows != expected_rows:
        raise ValueError(f"Expected {expected_rows} rows, got {data_rows}")
    if sheet.max_column != 9:
        raise ValueError(f"Expected 9 columns, got {sheet.max_column}")
    header = [sheet.cell(1, col).value for col in range(1, sheet.max_column + 1)]
    workbook.close()
    return data_rows, len(header), header


if __name__ == "__main__":
    total_rows, counts = build_workbook()
    verified_rows, verified_cols, verified_header = verify_workbook(total_rows)
    print(f"created={OUTPUT}")
    print(f"rows={verified_rows}")
    print(f"columns={verified_cols}")
    print("source_counts=" + "; ".join(f"{name}:{count}" for name, count in counts))
    print("header=" + ",".join(verified_header))
